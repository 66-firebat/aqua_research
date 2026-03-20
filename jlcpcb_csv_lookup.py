import csv
import time
import os
from typing import Optional
from jlcpcb_lookup_mod import search_components, JLCSEARCH_BASE

try:
    import openpyxl
    HAS_OPENPYXL: bool = True
except ImportError:
    openpyxl = None  # type: ignore
    HAS_OPENPYXL = False


def _get_file_type(filepath: str) -> str:
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".xlsx":
        return "xlsx"
    return "csv"


def lookup_packages(
    input_file: str,
    output_file: str,
    part_number_column: str = "Manf Part Number",
    delay: float = 0.5
) -> list[str]:
    """
    Read a CSV or XLSX file, look up each part number on JLCPCB, and add a Footprint column.

    Args:
        input_file: Path to input CSV/XLSX file
        output_file: Path to output CSV/XLSX file
        part_number_column: Name of the column containing part numbers
        delay: Seconds to wait between API calls (avoid rate limiting)

    Returns:
        List of package/footprint values for each row
    """
    file_type = _get_file_type(input_file)

    if file_type == "xlsx":
        rows, fieldnames = _read_xlsx(input_file, part_number_column)
    else:
        rows, fieldnames = _read_csv(input_file)

    if "Footprint" not in fieldnames:
        fieldnames.append("Footprint")

    footprints = []
    for row in rows:
        part_number = str(row.get(part_number_column, "")).strip()
        package = get_package_for_part(part_number, delay)
        footprints.append(package)
        row["Footprint"] = package

    if file_type == "xlsx":
        _write_xlsx(output_file, rows, fieldnames)
    else:
        _write_csv(output_file, rows, fieldnames)

    return footprints


def _read_csv(filepath: str) -> tuple[list[dict], list[str]]:
    with open(filepath, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames: list[str] = list(reader.fieldnames) if reader.fieldnames else []
        return rows, fieldnames


def _write_csv(filepath: str, rows: list[dict], fieldnames: list[str]) -> None:
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def _read_xlsx(filepath: str, part_number_column: str) -> tuple[list[dict], list[str]]:
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for XLSX support. Install with: pip install openpyxl")

    wb = openpyxl.load_workbook(filepath)  # type: ignore
    ws = wb.active
    assert ws is not None

    first_row = ws[1]
    headers: list[str] = [str(cell.value) if cell.value is not None else "" for cell in first_row]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))

    return rows, headers


def _write_xlsx(filepath: str, rows: list[dict], fieldnames: list[str]) -> None:
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required for XLSX support. Install with: pip install openpyxl")

    wb = openpyxl.Workbook()  # type: ignore
    ws = wb.active
    assert ws is not None

    ws.append(fieldnames)
    for row in rows:
        ws.append([row.get(field, "") for field in fieldnames])

    wb.save(filepath)


def get_package_for_part(part_number: str, delay: float = 0.5) -> str:
    """
    Search JLCPCB for a part number and return its package/footprint.

    Args:
        part_number: The manufacturer part number to search for
        delay: Seconds to wait after the request

    Returns:
        Package value or "bad_response" on failure
    """
    if not part_number:
        return "bad_response"

    try:
        results = search_components(part_number, limit=1)
        if results and len(results) > 0:
            package = results[0].get("package", "bad_response")
            time.sleep(delay)
            return package if package else "bad_response"
        return "bad_response"
    except Exception:
        time.sleep(delay)
        return "bad_response"


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Look up JLCPCB packages from CSV/XLSX")
    parser.add_argument("input_file", help="Input CSV or XLSX file path")
    parser.add_argument("output_file", help="Output CSV or XLSX file path")
    parser.add_argument(
        "--column",
        default="Manf Part Number",
        help="Column name containing part numbers (default: 'Manf Part Number')"
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=0.5,
        help="Delay between API calls in seconds (default: 0.5)"
    )

    args = parser.parse_args()

    print(f"Processing: {args.input_file}")
    print(f"Output: {args.output_file}")
    print(f"Column: {args.column}")

    footprints = lookup_packages(
        args.input_file,
        args.output_file,
        part_number_column=args.column,
        delay=args.delay
    )

    bad_count = sum(1 for fp in footprints if fp == "bad_response")
    print(f"\nDone! Processed {len(footprints)} parts.")
    print(f"Successful lookups: {len(footprints) - bad_count}")
    print(f"Bad responses: {bad_count}")
